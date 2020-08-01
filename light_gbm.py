'''
    Python Libraries Import
'''
import os
import numpy as np
import pandas as pd
import statsmodels.api as sm
from sklearn.linear_model import LogisticRegression, ElasticNet, ElasticNetCV
from sklearn.ensemble import RandomForestRegressor, AdaBoostRegressor
from sklearn.neural_network import MLPClassifier
from sklearn.svm import SVR
from sklearn.tree import DecisionTreeRegressor
from sklearn.discriminant_analysis import LinearDiscriminantAnalysis, QuadraticDiscriminantAnalysis
from sklearn.model_selection import train_test_split, cross_val_score, KFold, LeaveOneOut, cross_validate, cross_val_predict, GridSearchCV, RandomizedSearchCV
from sklearn.metrics import accuracy_score, confusion_matrix, classification_report, make_scorer
from openpyxl import load_workbook
import xlsxwriter
from scipy.stats import zscore
import seaborn as sns
from tqdm import tqdm
import eli5
from eli5.sklearn import PermutationImportance
import matplotlib.pyplot as plt
import warnings
from IPython.display import display
from sklearn.experimental import enable_hist_gradient_boosting
from sklearn.ensemble import HistGradientBoostingRegressor



'''
    Stub to read ABCD data from a specified path and to save the results of
    histogram gradient boosted trees to another specified file
'''
path =  '../ABCD 2.0/ABCDstudyNDA_RELEASE2/'
read_file = lambda file_name: pd.read_csv(file_name, delimiter='\t', skiprows=[1])
line = '-'*100
workbook = xlsxwriter.Workbook('hgbr_results.xlsx')
workbook.close()

save = pd.DataFrame(columns=['Dependent', 'Analysis Type', 'Model', 'Parameters', 'MAE (in sigma)', 'Explained Variance', 'R^2'])
global row
row = 1



'''
    Code to carry out Regression Analysis using Histogram Gradient Boosting Regressor and Permutation Feature Importance
'''
def hgbr_predict(X, y, fname='Untitled'):
    global row
    k_fold=10
    learning_rates = [1, 0.5, 0.25, 0.1, 0.05, 0.01]
    max_depth = [10, 20, 30]
    l2_regularization = [0, 0.25, 0.5, 0.75, 1]

    random_grid = {'max_depth': max_depth,
                   'l2_regularization':l2_regularization,
                   'learning_rate':learning_rates}
    gb = HistGradientBoostingRegressor(max_iter=100, max_leaf_nodes=50)
    gb_fit = GridSearchCV(estimator=gb, param_grid=random_grid, cv=10, verbose=2, scoring='r2', n_jobs=-1)
    gb_fit.fit(X, y)
    best_estimator = gb_fit.best_estimator_
    best_params = gb_fit.best_params_

    print(best_estimator)
    scores = cross_validate(best_estimator, X, y, cv=k_fold, scoring=('r2', 'neg_mean_absolute_error', 'explained_variance'), return_train_score=True, return_estimator=True, n_jobs=-1)
    mae = ("%.4f sigma ± %.4f sigma")%(-scores['test_neg_mean_absolute_error'].mean()/y.std(), scores['test_neg_mean_absolute_error'].std()/y.std())
    exp_var = ("%.4f ± %.4f")%(scores['test_explained_variance'].mean(), scores['test_explained_variance'].std())
    r2 = ("%.4f ± %.4f")%(scores['test_r2'].mean(), scores['test_r2'].std())
    print('MAE                : %s'%(mae))
    print('Explained Variance : %s'%(exp_var))
    print('R^2                : %s'%(r2))

    dep, typ1, typ2 = fname.split('_')
    save.loc[row, :] = [dep, typ1+'_'+typ2, 'Light Gradient Boosting Regressor', best_params, mae, exp_var, r2]
    row = row + 1

    perm = PermutationImportance(best_estimator, cv=k_fold, n_iter=100).fit(X, y)
    perm_df = eli5.explain_weights(perm, feature_names = X.columns.tolist())
    perm_df = pd.DataFrame()
    perm_df['feature'] = X.columns.tolist()
    perm_df['feature_imp'] = perm.feature_importances_
    perm_df['feature_imp_std'] = perm.feature_importances_std_
    perm_df = perm_df.sort_values(by=['feature_imp'], ascending=False).set_index('feature')

    file = 'hgbr_results.xlsx'
    book = load_workbook(file)
    writer = pd.ExcelWriter(file, engine='openpyxl')
    writer.book = book
    perm_df.to_excel(writer, index=True, sheet_name='weights_'+fname)
    writer.save()

def predict(X, y, fname='Untitled'):
    hgbr_predict(X, y, fname)



'''
    Code to prepare and run ABCD data for analysis
'''
def run(dep, dep_cov, title):
    dependent = dep
    file = 'abcd_cbcls01.txt'
    features = ['subjectkey',dependent]
    depression = pd.read_csv(path + file, delimiter='\t', skiprows=[1], low_memory=False)
    depression = depression[depression['eventname'] == 'baseline_year_1_arm_1']
    depression = depression[features].dropna()
    cov = read_file(path + 'abcd_asrs01.txt')[['subjectkey', dep_cov]].dropna()

    index = pd.read_pickle('../data/no_twins_triplets').drop('rel_family_id', axis=1)
    struct = pd.read_pickle('../data/structural').reset_index()
    data = struct.merge(depression, how='inner', on='subjectkey', validate='1:1')
    data = data.merge(index, how='inner', on='subjectkey', validate='1:1')
    file = path + 'pdem02.txt'
    features = ['subjectkey', 'demo_brthdat_v2']
    df = pd.read_csv(file, delimiter='\t', skiprows=[1])[features].dropna()
    data = pd.merge(data, df, on='subjectkey', validate='1:1').set_index('subjectkey')
    X = data.drop([dependent], axis=1)
    for col in list(X.dtypes[X.dtypes == 'category'].keys()):
        X[col] = X[col].factorize()[0]
    y = data[dependent]
    print(line)
    print('sMRI Results Without Covariates')
    print(data.shape)
    predict(X, y, title+'_sMRI_woCov')
    print(line)

    index = pd.read_pickle('../data/no_twins_triplets').drop('rel_family_id', axis=1)
    func = pd.read_pickle('../data/functional').reset_index()
    data = func.merge(depression, how='inner', on='subjectkey', validate='1:1')
    data = data.merge(index, how='inner', on='subjectkey', validate='1:1')
    features = ['subjectkey', 'demo_brthdat_v2']
    df = pd.read_csv(file, delimiter='\t', skiprows=[1])[features].dropna()
    data = pd.merge(data, df, on='subjectkey', validate='1:1').set_index('subjectkey')
    X = data.drop([dependent], axis=1)
    for col in list(X.dtypes[X.dtypes == 'category'].keys()):
        X[col] = X[col].factorize()[0]
    y = data[dependent]
    print('fMRI Results Without Covariates')
    print(data.shape)
    predict(X, y, title+'_fMRI_woCov')
    print(line)

    index = pd.read_pickle('../data/no_twins_triplets')
    covariates = pd.read_pickle('../data/covariates').reset_index()
    cov = cov.set_index('subjectkey').apply(zscore).reset_index()
    covariates = covariates.merge(cov, how='inner', on='subjectkey', validate='1:1')
    struct = pd.read_pickle('../data/structural').reset_index()
    data = struct.merge(depression, how='inner', on='subjectkey', validate='1:1')
    data = data.merge(covariates, how='inner', on='subjectkey', validate='1:1')
    data = data.merge(index, how='inner', on='subjectkey', validate='1:1').set_index('subjectkey')
    X = data.drop(dependent, axis=1)
    for col in list(X.dtypes[X.dtypes == 'category'].keys()):
        X[col] = X[col].factorize()[0]
    y = data[dependent]
    print('sMRI Results With Covariates')
    print(data.shape)
    predict(X, y, title+'_sMRI_wCov')
    print(line)

    index = pd.read_pickle('../data/no_twins_triplets')
    func = pd.read_pickle('../data/functional').reset_index()
    data = func.merge(depression, how='inner', on='subjectkey', validate='1:1')
    data = data.merge(covariates, how='inner', on='subjectkey', validate='1:1')
    data = data.merge(index, how='inner', on='subjectkey', validate='1:1').set_index('subjectkey')
    X = data.drop(dependent, axis=1)
    for col in list(X.dtypes[X.dtypes == 'category'].keys()):
        X[col] = X[col].factorize()[0]
    y = data[dependent]
    print('fMRI Results With Covariates')
    print(data.shape)
    predict(X, y, title+'_fMRI_wCov')
    print(line)

    index = pd.read_pickle('../data/no_twins_triplets')
    covariates = pd.read_pickle('../data/covariates').reset_index()
    cov = cov.set_index('subjectkey').apply(zscore).reset_index()
    covariates = covariates.merge(cov, how='inner', on='subjectkey', validate='1:1')
    data = covariates.merge(depression, how='inner', on='subjectkey', validate='1:1')
    data = data.merge(index, how='inner', on='subjectkey', validate='1:1').set_index('subjectkey')
    X = data.drop(dependent, axis=1)
    for col in list(X.dtypes[X.dtypes == 'category'].keys()):
        X[col] = X[col].factorize()[0]
    y = data[dependent]
    print('Only Covariates')
    print(data.shape)
    predict(X, y, title+'_only_Cov')
    print(line)

    index = pd.read_pickle('../data/no_twins_triplets')
    data = struct.merge(depression, how='inner', on='subjectkey', validate='1:1')
    data = data.merge(func, how='inner', on='subjectkey', validate='1:1')
    data = data.merge(covariates, how='inner', on='subjectkey', validate='1:1')
    data = data.merge(index, how='inner', on='subjectkey', validate='1:1').set_index('subjectkey')
    X = data.drop(dependent, axis=1)
    for col in list(X.dtypes[X.dtypes == 'category'].keys()):
        X[col] = X[col].factorize()[0]
    y = data[dependent]
    print('sMRI & fMRI Results With Covariates')
    print(data.shape)
    predict(X, y, title+'_sMRI+fMRI_wCov')
    print(line)

run('cbcl_scr_dsm5_depress_t', 'asr_scr_anxdep_t', 'Depression') # CBCL Depression
run('cbcl_scr_syn_totprob_t', 'asr_scr_totprob_t', 'Total Problems') # CBCL Total Problems
run('cbcl_scr_syn_internal_t', 'asr_scr_internal_t', 'Internal') # CBCL Internalizing
run('cbcl_scr_syn_external_t', 'asr_scr_external_t', 'External') # CBCL Externalizing

file = 'hgbr_results.xlsx'
book = load_workbook(file)
book.remove_sheet(book.get_sheet_by_name('Sheet1'))
writer = pd.ExcelWriter(file, engine='openpyxl')
writer.book = book
save.to_excel(writer, sheet_name='performance')
writer.save()
writer.close()
